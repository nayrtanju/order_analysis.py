import streamlit as st
import tempfile
import os
import traceback
import re
from io import BytesIO

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage


st.set_page_config(
    page_title="Vehicle Order Analysis Tool",
    layout="wide"
)

st.title("Vehicle Order Analysis Tool")


try:
    from order_analysis import (
        read_xlsx_numeric,
        angular_resample,
        order_map,
        extract_order_vs_rpm
    )
except Exception:
    st.error("order_analysis.py yüklenirken hata oluştu")
    st.code(traceback.format_exc())
    st.stop()


MAX_FILE_SIZE_MB = 500
MAX_ROWS = 3000000


TARGETS = {
    "Diesel": {
        "Front Axle": {
            "rpm": np.array([1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500]),
            "amp": np.array([2.5, 2.5, 2.5, 7.5, 7.5, 7.5, 7.5, 7.5])
        },
        "Rear Axle": {
            "rpm": np.array([1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500]),
            "amp": np.array([2.5, 2.5, 2.5, 7.5, 7.5, 7.5, 7.5, 7.5])
        }
    },
    "Gasoline": {
        "Front Axle": {
            "rpm": np.array([1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500]),
            "amp": np.array([2.5, 2.5, 2.5, 6.25, 10.0, 10.0, 10.0, 10.0])
        },
        "Rear Axle": {
            "rpm": np.array([5.0, 5.0, 5.0, 10.0, 12.5, 12.5, 12.5, 12.5]),
            "amp": np.array([5.0, 5.0, 5.0, 10.0, 12.5, 12.5, 12.5, 12.5])
        }
    }
}

# Correct Gasoline Rear Axle RPM axis
TARGETS["Gasoline"]["Rear Axle"]["rpm"] = np.array(
    [1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500]
)


TARGET_ORDERS = [10.0, 20.0]


def load_measurement_file(uploaded_file):
    file_extension = uploaded_file.name.split(".")[-1].lower()

    if uploaded_file.size > MAX_FILE_SIZE_MB * 1024 * 1024:
        st.error(f"File exceeds maximum allowed size: {MAX_FILE_SIZE_MB} MB.")
        st.stop()

    if file_extension == "xlsx":
        temp_file = None

        try:
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".xlsx"
            ) as tmp:
                tmp.write(uploaded_file.read())
                temp_file = tmp.name

            headers, data = read_xlsx_numeric(temp_file)

        finally:
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass

    elif file_extension == "csv":
        try:
            df = pd.read_csv(
                uploaded_file,
                sep=None,
                engine="python"
            )

            headers = list(df.columns)
            data = df.to_numpy(dtype=float)

        except Exception:
            st.error("CSV file could not be read. Please check delimiter and numeric data format.")
            st.code(traceback.format_exc())
            st.stop()

    else:
        st.error("Unsupported file format. Please upload .xlsx or .csv file.")
        st.stop()

    if data.ndim != 2 or data.shape[1] < 5:
        st.error(
            "Measurement file must contain at least 5 columns: Time, ChA, ChB, ChC, RPM."
        )
        st.stop()

    if data.shape[0] > MAX_ROWS:
        st.error(f"Dataset exceeds maximum row limit: {MAX_ROWS} rows.")
        st.stop()

    if data.shape[0] < 10:
        st.error("Dataset is too short for order analysis.")
        st.stop()

    if not np.all(np.isfinite(data[:, :5])):
        st.error("Dataset contains NaN or non-numeric values in the first 5 columns.")
        st.stop()

    time = data[:, 0]
    rpm = data[:, 4]

    if not np.all(np.diff(time) > 0):
        st.error("Time column must be strictly increasing.")
        st.stop()

    if np.any(rpm <= 0):
        st.error("RPM column must contain only positive values.")
        st.stop()

    return headers, data


def format_comparison_sheet(writer, sheet_name):
    ws = writer.book[sheet_name]

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    status_col = None

    for cell in ws[1]:
        if cell.value == "Status":
            status_col = cell.column
            break

    if status_col is not None:
        for row in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=status_col)

            if status_cell.value == "PASS":
                fill = green_fill
            elif status_cell.value == "FAIL":
                fill = red_fill
            else:
                fill = None

            if fill is not None:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 22


def format_curve_sheet(writer, sheet_name):
    ws = writer.book[sheet_name]

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 16


def create_curve_plot_png(
    curve_df,
    order_value,
    vin_number,
    fuel_type,
    axle_type
):
    fig, ax = plt.subplots(figsize=(14, 8))

    ax.plot(curve_df["RPM"], curve_df["ChA"], label="ChA", linewidth=2)
    ax.plot(curve_df["RPM"], curve_df["ChB"], label="ChB", linewidth=2)
    ax.plot(curve_df["RPM"], curve_df["ChC"], label="ChC", linewidth=2)

    ax.plot(
        curve_df["RPM"],
        curve_df["Target"],
        label="Target Curve",
        color="red",
        linewidth=5
    )

    ax.set_title(
        f"{int(order_value)}. Order vs RPM | VIN: {vin_number} | {fuel_type} | {axle_type}",
        fontsize=16
    )

    ax.set_xlabel("RPM", fontsize=13)
    ax.set_ylabel(f"{int(order_value)}. Order Amplitude [m/s²]", fontsize=13)

    ax.grid(True, alpha=0.3)
    ax.legend(loc="upper right", fontsize=12)

    rpm_min = min(1000, float(curve_df["RPM"].min()))
    rpm_max = max(4500, float(curve_df["RPM"].max()))
    ax.set_xlim(rpm_min, rpm_max)

    fig.tight_layout()

    img_buffer = BytesIO()
    fig.savefig(
        img_buffer,
        format="png",
        dpi=180,
        bbox_inches="tight"
    )
    plt.close(fig)

    img_buffer.seek(0)
    return img_buffer


def add_png_plot_to_sheet(
    writer,
    sheet_name,
    curve_df,
    order_value,
    vin_number,
    fuel_type,
    axle_type
):
    ws = writer.book[sheet_name]

    img_buffer = create_curve_plot_png(
        curve_df=curve_df,
        order_value=order_value,
        vin_number=vin_number,
        fuel_type=fuel_type,
        axle_type=axle_type
    )

    img = XLImage(img_buffer)
    img.width = 900
    img.height = 520

    ws.add_image(img, "G2")


def make_excel_report(vehicle_info, results_by_order, curves_by_order):
    output = BytesIO()

    vin_number = vehicle_info["VIN"]
    fuel_type = vehicle_info["Fuel Type"]
    axle_type = vehicle_info["Axle Type"]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame([vehicle_info]).to_excel(
            writer,
            sheet_name="Vehicle Info",
            index=False
        )

        for order_value, result_df in results_by_order.items():
            sheet_name = f"{int(order_value)} Order Comparison"
            sheet_name = sheet_name[:31]

            result_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            format_comparison_sheet(writer, sheet_name)

        for order_value, curve_df in curves_by_order.items():
            sheet_name = f"{int(order_value)} Order Curves"
            sheet_name = sheet_name[:31]

            curve_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            format_curve_sheet(writer, sheet_name)

            add_png_plot_to_sheet(
                writer=writer,
                sheet_name=sheet_name,
                curve_df=curve_df,
                order_value=order_value,
                vin_number=vin_number,
                fuel_type=fuel_type,
                axle_type=axle_type
            )

    output.seek(0)
    return output


def analyze_target_order(
    order_value,
    time,
    rpm,
    channels,
    samples_per_rev,
    revs_per_block,
    overlap,
    max_order,
    order_width,
    rpm_step,
    cal_factor,
    target_rpm,
    target_amp
):
    channel_curves = {}
    peak_results = []

    for name, sig in channels.items():

        theta_u, x_u, rpm_u = angular_resample(
            time,
            rpm,
            sig,
            samples_per_rev=samples_per_rev
        )

        orders, rpms, spec = order_map(
            theta_u,
            x_u,
            rpm_u,
            samples_per_rev=samples_per_rev,
            revs_per_block=revs_per_block,
            overlap=overlap,
            max_order=max_order
        )

        rpm_sorted, amp_sorted = extract_order_vs_rpm(
            orders,
            rpms,
            spec,
            target_order=order_value,
            width=order_width,
            rpm_step=rpm_step,
            smooth=True
        )

        amp_sorted = amp_sorted * cal_factor

        channel_curves[name] = {
            "rpm": rpm_sorted,
            "amp": amp_sorted
        }

        peak_idx = np.argmax(amp_sorted)
        peak_rpm = float(rpm_sorted[peak_idx])
        peak_amp = float(amp_sorted[peak_idx])

        target_at_peak = float(
            np.interp(
                peak_rpm,
                target_rpm,
                target_amp
            )
        )

        margin = peak_amp - target_at_peak
        margin_percent = (
            margin / target_at_peak * 100.0
            if target_at_peak > 0
            else np.nan
        )

        status = "PASS" if peak_amp <= target_at_peak else "FAIL"

        peak_results.append({
            "Order": order_value,
            "Channel": name,
            "Peak RPM": peak_rpm,
            "Peak Amplitude [m/s²]": peak_amp,
            "Target at Peak RPM [m/s²]": target_at_peak,
            "Margin [m/s²]": margin,
            "Margin [%]": margin_percent,
            "Status": status
        })

    result_df = pd.DataFrame(peak_results)

    curve_df = pd.DataFrame()
    base_rpm = None

    for name, curve in channel_curves.items():
        if base_rpm is None:
            base_rpm = curve["rpm"]
            curve_df["RPM"] = base_rpm

        curve_df[name] = np.interp(
            base_rpm,
            curve["rpm"],
            curve["amp"]
        )

    curve_df["Target"] = np.interp(
        curve_df["RPM"],
        target_rpm,
        target_amp
    )

    return channel_curves, result_df, curve_df


def plot_order_comparison(
    order_value,
    channel_curves,
    target_rpm,
    target_amp,
    vin_number,
    fuel_type,
    axle_type
):
    fig, ax = plt.subplots(figsize=(12, 7))

    for name, curve in channel_curves.items():
        ax.plot(
            curve["rpm"],
            curve["amp"],
            label=name
        )

    ax.plot(
        target_rpm,
        target_amp,
        color="red",
        linewidth=4,
        label="Target Curve"
    )

    ax.set_xlabel("RPM")
    ax.set_ylabel(f"{int(order_value)}. Order Amplitude [m/s²]")
    ax.set_title(
        f"{int(order_value)}. Order vs RPM | VIN: {vin_number} | {fuel_type} | {axle_type}"
    )

    ax.grid(True, alpha=0.3)
    ax.legend()

    return fig


st.subheader("Vehicle Information")

col1, col2, col3 = st.columns(3)

with col1:
    vin_number = st.text_input(
        "VIN Number",
        placeholder="Enter 17-character VIN",
        max_chars=17
    ).upper().strip()

vin_valid = bool(
    re.fullmatch(r"[A-Z0-9]{17}", vin_number)
)

with col2:
    fuel_type = st.selectbox(
        "Fuel Type",
        ["Select fuel type", "Diesel", "Gasoline"],
        disabled=not vin_valid
    )

with col3:
    axle_type = st.selectbox(
        "Axle Type",
        ["Select axle type", "Front Axle", "Rear Axle"],
        disabled=not vin_valid
    )

if vin_number and not vin_valid:
    st.error(
        "VIN must be exactly 17 characters and contain only letters and numbers."
    )


st.subheader("Measurement Data")

uploaded_file = st.file_uploader(
    "Upload Measurement File",
    type=["xlsx", "csv"],
    disabled=not vin_valid,
    help="Supported formats: .xlsx and .csv"
)


can_continue = (
    vin_valid
    and fuel_type != "Select fuel type"
    and axle_type != "Select axle type"
    and uploaded_file is not None
)

if not can_continue:
    if not vin_valid:
        st.warning(
            "Please enter a valid 17-character VIN before selecting fuel type and uploading data."
        )
    else:
        st.warning(
            "Please select fuel type, select axle type, and upload measurement file."
        )
    st.stop()


target_rpm = TARGETS[fuel_type][axle_type]["rpm"]
target_amp = TARGETS[fuel_type][axle_type]["amp"]

st.success("Vehicle information and measurement file are ready for analysis.")

info_cols = st.columns(3)
info_cols[0].metric("VIN", vin_number)
info_cols[1].metric("Fuel Type", fuel_type)
info_cols[2].metric("Axle Type", axle_type)


st.subheader("Analysis Settings")

samples_per_rev = 512
revs_per_block = 8
overlap = 0.75
rpm_step = 10
cal_factor = 1.0

with st.expander("Advanced Settings", expanded=False):

    selected_channel = st.selectbox(
        "Order Map Channel",
        ["ChA", "ChB", "ChC"]
    )

    max_order = st.slider(
        "Max order",
        5,
        80,
        30
    )

    order_width = st.number_input(
        "Order width",
        min_value=0.05,
        max_value=2.0,
        value=0.15,
        step=0.05
    )


if st.button("Run Order Analysis", type="primary"):

    try:
        headers, data = load_measurement_file(uploaded_file)

        time = data[:, 0]
        rpm = data[:, 4]

        channels = {
            "ChA": data[:, 1],
            "ChB": data[:, 2],
            "ChC": data[:, 3],
        }

        with st.spinner("Order analysis is running..."):

            curves_by_order = {}
            results_by_order = {}
            raw_curves_by_order = {}

            for order_value in TARGET_ORDERS:

                channel_curves, result_df, curve_df = analyze_target_order(
                    order_value=order_value,
                    time=time,
                    rpm=rpm,
                    channels=channels,
                    samples_per_rev=samples_per_rev,
                    revs_per_block=revs_per_block,
                    overlap=overlap,
                    max_order=max_order,
                    order_width=order_width,
                    rpm_step=rpm_step,
                    cal_factor=cal_factor,
                    target_rpm=target_rpm,
                    target_amp=target_amp
                )

                curves_by_order[order_value] = channel_curves
                results_by_order[order_value] = result_df
                raw_curves_by_order[order_value] = curve_df

            overall_status = "PASS"

            for result_df in results_by_order.values():
                if not (result_df["Status"] == "PASS").all():
                    overall_status = "FAIL"

            tab1, tab2, tab3, tab4 = st.tabs(
                [
                    "10th Order Target Comparison",
                    "20th Order Target Comparison",
                    "Order Map",
                    "Raw Results"
                ]
            )

            for tab, order_value in zip([tab1, tab2], TARGET_ORDERS):

                with tab:
                    result_df = results_by_order[order_value]
                    channel_curves = curves_by_order[order_value]

                    order_status = (
                        "PASS"
                        if (result_df["Status"] == "PASS").all()
                        else "FAIL"
                    )

                    st.subheader(f"{int(order_value)}th Order Result Summary")

                    kpi1, kpi2, kpi3, kpi4 = st.columns(4)

                    kpi1.metric(
                        "Peak ChA",
                        f"{result_df.loc[result_df['Channel'] == 'ChA', 'Peak Amplitude [m/s²]'].iloc[0]:.2f} m/s²"
                    )

                    kpi2.metric(
                        "Peak ChB",
                        f"{result_df.loc[result_df['Channel'] == 'ChB', 'Peak Amplitude [m/s²]'].iloc[0]:.2f} m/s²"
                    )

                    kpi3.metric(
                        "Peak ChC",
                        f"{result_df.loc[result_df['Channel'] == 'ChC', 'Peak Amplitude [m/s²]'].iloc[0]:.2f} m/s²"
                    )

                    kpi4.metric(
                        f"{int(order_value)}th Order Assessment",
                        order_status
                    )

                    st.subheader(f"{int(order_value)}th Order vs RPM with Target Curve")

                    fig_cmp = plot_order_comparison(
                        order_value=order_value,
                        channel_curves=channel_curves,
                        target_rpm=target_rpm,
                        target_amp=target_amp,
                        vin_number=vin_number,
                        fuel_type=fuel_type,
                        axle_type=axle_type
                    )

                    st.pyplot(fig_cmp)

                    st.subheader(f"{int(order_value)}th Order Target Compliance")

                    st.dataframe(
                        result_df,
                        use_container_width=True
                    )

                    if order_status == "PASS":
                        st.success(f"{int(order_value)}th Order Assessment: PASS")
                    else:
                        st.error(f"{int(order_value)}th Order Assessment: FAIL")

                    png_buffer = BytesIO()
                    fig_cmp.savefig(
                        png_buffer,
                        format="png",
                        dpi=200,
                        bbox_inches="tight"
                    )
                    png_buffer.seek(0)

                    st.download_button(
                        label=f"Download {int(order_value)}th Order Target Comparison PNG",
                        data=png_buffer,
                        file_name=f"{vin_number}_{int(order_value)}th_order_target_comparison.png",
                        mime="image/png"
                    )

            with tab3:

                st.subheader(f"Order Map - {selected_channel}")

                sig = channels[selected_channel]

                theta_u, x_u, rpm_u = angular_resample(
                    time,
                    rpm,
                    sig,
                    samples_per_rev=samples_per_rev
                )

                orders, rpms, spec = order_map(
                    theta_u,
                    x_u,
                    rpm_u,
                    samples_per_rev=samples_per_rev,
                    revs_per_block=revs_per_block,
                    overlap=overlap,
                    max_order=max_order
                )

                idx = np.argsort(rpms)
                r = rpms[idx]
                s = spec[idx]

                db = 20 * np.log10(
                    np.maximum(s * cal_factor, 1e-12)
                )

                fig, ax = plt.subplots(figsize=(12, 7))

                im = ax.imshow(
                    db,
                    aspect="auto",
                    origin="lower",
                    extent=[orders[0], orders[-1], r[0], r[-1]],
                    interpolation="nearest",
                    cmap="jet"
                )

                fig.colorbar(
                    im,
                    ax=ax,
                    label="Amplitude [dB re 1 m/s²]"
                )

                ax.set_xlabel("Order")
                ax.set_ylabel("RPM")
                ax.set_title(
                    f"Order Map - {selected_channel} | VIN: {vin_number} | {fuel_type} | {axle_type}"
                )

                st.pyplot(fig)

            with tab4:

                st.subheader("10th Order Raw Curve Data")

                st.dataframe(
                    raw_curves_by_order[10.0],
                    use_container_width=True
                )

                st.subheader("20th Order Raw Curve Data")

                st.dataframe(
                    raw_curves_by_order[20.0],
                    use_container_width=True
                )

                st.subheader("Overall Assessment")

                if overall_status == "PASS":
                    st.success("Overall Assessment: PASS")
                else:
                    st.error("Overall Assessment: FAIL")

                vehicle_info = {
                    "VIN": vin_number,
                    "Fuel Type": fuel_type,
                    "Axle Type": axle_type,
                    "Target Orders": "10, 20",
                    "Order Width": order_width,
                    "RPM Step": rpm_step,
                    "Samples per Rev": samples_per_rev,
                    "Revs per Block": revs_per_block,
                    "Overlap": overlap,
                    "Calibration Factor": cal_factor,
                    "Max Order": max_order,
                    "Overall Assessment": overall_status
                }

                excel_report = make_excel_report(
                    vehicle_info,
                    results_by_order,
                    raw_curves_by_order
                )

                st.download_button(
                    label="Download Excel Report",
                    data=excel_report,
                    file_name=f"{vin_number}_order_analysis_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception:
        st.error("Uygulama çalışırken hata oluştu")
        st.code(traceback.format_exc())
